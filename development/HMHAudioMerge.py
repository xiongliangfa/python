import speech_recognition as sr
import subprocess
import wave,os
from tkinter import *
from tkinter import messagebox
import tkinter as tk
from tkinter.filedialog import askopenfilename
import openpyxl

# global deceleration
sox = 'C:/Program Files (x86)/sox-14-4-2/sox.exe'
file_list = []
win = Tk()
win.title('Audio Merge tool V2.1 - HMH')
mystring = StringVar()
audioLabel = StringVar()
countLabel = StringVar()
excelFile = StringVar()

# function to return output audio name
def audOutput(outputAudio):
    if os.path.isfile(excelFile.get()):
        AudioFName = sheet_obj.cell(0,i).value
        retAudio = AudioFName + ".wav"
        print(i)
        i+=1
    else:
        retAudio = outputAudio
    return retAudio

#func to merge the audio
def mergeAudio(aud1,aud2,outputAudio):
    infiles = [aud1, aud2]
    outfile = outputAudio
    data = []
    for infile in infiles:
        w = wave.open(infile, 'rb')
        data.append( [w.getparams(), w.readframes(w.getnframes())])
        w.close()
    output = wave.open(outfile, 'wb')
    output.setparams(data[0][0])
    output.writeframes(data[0][1])
    output.writeframes(data[1][1])
    output.close()

# func for converting audio to text
def SpeechToText(audioName,k): 
    r = sr.Recognizer()
    harvard = sr.AudioFile(audioName)
    with harvard as source:
        audio = r.record(source)
        type(audio)
        try:
            convertedText = r.recognize_google(audio, language="en-US")
            # print(convertedText)
        except:
            convertedText = 'test' + str(k)
        outAudioName = convertedText + '.wav'
        return outAudioName

# adding files to array
def addToArray(filePath):
    file_list.clear()
    directory = os.listdir(filePath)
    os.chdir(filePath)
    for file in directory:
        file_list.append(file)

#  adding silence
def addSilence(aud1):
    infiles = [aud1, 'silence\\silence.wav']
    outfile = aud1
    data = []
    for infile in infiles:
        w = wave.open(infile, 'rb')
        data.append( [w.getparams(), w.readframes(w.getnframes())] )
        w.close()
    output = wave.open(outfile, 'wb')
    output.setparams(data[0][0])
    output.writeframes(data[0][1])
    output.writeframes(data[1][1])
    output.close()

# mp3 to wav function
def toWav(filePath):
    for i in range(0,len(file_list)):
        audioName = os.path.splitext(file_list[i])[0]
        if not file_list[i] == 'process':
            infile = filePath + '\\' + audioName + '.mp3'
            outfile = filePath + '\\process\\' + audioName + '.wav'
            extra = '--rate 16k --bits 16 --channels 1'
            command = """{0} "{1}" {2} "{3}" """.format(sox,infile,extra,outfile)
            subprocess.call(command)

# wav to mp3 function
def tomp3(filePath):
    for i in range(0,len(file_list)):
        if not file_list[i] == 'silence' or file_list[i] == 'log.txt':
            audioName = os.path.splitext(file_list[i])[0]
            infile = filePath + '\\process\\' + audioName + '.wav'
            outfile = filePath + '\\' + audioName + '.mp3'
            command = """{0} "{1}" "{2}" """.format(sox,infile,outfile)
            subprocess.call(command)

# triggering function
def trigger():
    try:
        if excelFile.get() == None:
            wb_obj = openpyxl.load_workbook(excelFile.get())
            sheet_obj = wb_obj.active
        filePath = mystring.get()  
        addToArray(filePath)
        toWav(filePath)
        os.chdir(filePath + "\\process")
        addToArray(filePath + "\\process")
        file = open('log.txt','w')
        
        for i in range(0,len(file_list),2):
            k=i/2+1
            if not file_list[i] == 'silence':
                win.update_idletasks()
                # addSilence(file_list[i])
                if os.path.isfile(excelFile.get()):
                    AudioFName = sheet_obj.cell(k,1).value
                    if not sheet.cell(k, 1).value == xlrd.empty_cell.value:
                        outputAudio = AudioFName + ".wav"
                    else:
                        messagebox.showerror("Failed", "Entry in excel cannot be found.")
                else:
                    outputAudio = SpeechToText(file_list[i],i)
                mergeAudio(file_list[i],file_list[i+1],outputAudio)
                file.write(file_list[i] + ' + ' + file_list[i+1] + ' --> ' + outputAudio + '\n')
        file.close()
        addToArray(filePath + "\\process")
        tomp3(filePath)
        messagebox.showinfo("Processed", "Files Processed")
    except:
        messagebox.showerror("Failed", "Process Failed")
        pass

# fetchfile
def fetchExcel():
    fileExcel = askopenfilename()
    excelFile.set(fileExcel)
    return fileExcel

if __name__ == "__main__":
    audioLabel.set('Audio to be generated')
    countLabel.set('0')
    Label(win, text="").pack()  #label
    Label(win, text="Enter Audio Path", justify=LEFT).pack() #label
    Entry(win, textvariable = mystring,width=50).pack() #textblock
    proceed = Button(win, text="Proceed", command=trigger) #button
    proceed.pack()
    Entry(win, textvariable = excelFile, width = 50).pack()
    fetchButton = Button(win, text = "Browse", command = fetchExcel)
    fetchButton.pack()
    win.geometry("350x150+500+250")
    win.mainloop()